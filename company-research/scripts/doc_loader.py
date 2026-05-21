#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文档加载器 - PDF转图片（PPT来源）或OCR识别（Word来源），其他文档提文本
自包含脚本，不依赖外部项目代码

特性：
- 支持PDF/DOCX/TXT/MD/CSV/JSON/XLSX
- PDF根据类型自动选择处理方式：
  * PPT转PDF（图片型）→ 转图片，AI直接读取
  * Word转PDF（文字型）→ OCR识别文字
- 不输出任何文本内容到终端，只输出文件清单
- 首次使用自动安装所需依赖

依赖（仅2个）：pymupdf, pillow
  pip install pymupdf pillow

用法:
  python doc_loader.py <资料目录>
  python doc_loader.py <资料目录> --output-dir ./cache
"""
import sys
import os
import json
import argparse
import hashlib
import subprocess
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")


# ==================== 依赖自动安装 ====================
def check_and_install_dependencies():
    required = {'pymupdf': 'pymupdf', 'PIL': 'pillow', 'pytesseract': 'pytesseract', 'openpyxl': 'openpyxl', 'docx': 'python-docx'}
    missing = []
    for module, package in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(package)

    if missing:
        print(f"  正在安装依赖: {', '.join(missing)} ...", file=sys.stderr)
        try:
            subprocess.run([sys.executable, '-m', 'pip', 'install'] + missing,
                         capture_output=True, check=True)
            print(f"  依赖安装成功", file=sys.stderr)
        except Exception as e:
            print(f"  警告: 依赖安装失败. {e}", file=sys.stderr)
            print(f"  请手动: pip install {' '.join(missing)}", file=sys.stderr)
            print(f"  注意: pytesseract还需安装Tesseract OCR引擎", file=sys.stderr)

    
check_and_install_dependencies()


SUPPORTED_EXTENSIONS = {".pdf", ".doc", ".docx", ".txt", ".md", ".csv", ".json", ".xls", ".xlsx"}


# ==================== 文本加载器 ====================

def load_text_file(path: Path) -> str:
    for enc in ("utf-8", "gbk", "gb2312", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    return f"[无法解码: {path.name}]"


def load_docx_file(path: Path) -> str:
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)
    except ImportError:
        return f"[缺少python-docx: {path.name}]"
    except Exception as e:
        return f"[无法读取: {path.name}, {e}]"


def load_csv_file(path: Path) -> str:
    try:
        import pandas as pd
        df = pd.read_csv(path)
        return df.to_string()
    except ImportError:
        try:
            rows = []
            for enc in ("utf-8", "gbk"):
                try:
                    with open(path, "r", encoding=enc) as f:
                        for line in f:
                            rows.append(line.rstrip())
                    return "\n".join(rows[:200])
                except (UnicodeDecodeError, UnicodeError):
                    continue
            return f"[无法解码: {path.name}]"
        except Exception as e:
            return f"[无法读取: {path.name}, {e}]"


def _classify_sheet(sheet_name: str, header_row: list) -> str:
    """根据 sheet 名和表头推测 sheet 类型"""
    sn = sheet_name.lower()
    if any(k in sn for k in ("利润", "损益", "income", "p&l", "pnl", "profit")):
        return "利润表"
    if any(k in sn for k in ("资产负债", "balance", "bs")):
        return "资产负债表"
    if any(k in sn for k in ("现金流量", "cashflow", "cash_flow", "cf")):
        return "现金流量表"
    if any(k in sn for k in ("估值", "valuation", "dcf", "model", "模型")):
        return "估值模型"
    header_text = " ".join(str(c).lower() for c in header_row if c)
    if any(k in header_text for k in ("revenue", "收入", "营收", "营业额")):
        if any(k in header_text for k in ("cost", "成本", "费用")):
            return "利润表"
    if any(k in header_text for k in ("资产", "asset", "负债", "liability")):
        return "资产负债表"
    if any(k in header_text for k in ("现金流", "cash flow")):
        return "现金流量表"
    return "其他"


def _extract_key_metrics(sheet_type: str, rows: list, col_count: int) -> list:
    """从财务 sheet 提取核心指标行"""
    metric_keywords = {
        "利润表": ["营业收入", "营业成本", "毛利", "毛利率", "营业利润", "净利润", "净利率",
                    "营收", "收入", "revenue", "gross profit", "net income", "ebitda"],
        "资产负债表": ["总资产", "总负债", "净资产", "所有者权益", "现金", "应收",
                      "total assets", "total liabilities", "equity"],
        "现金流量表": ["经营活动", "投资活动", "筹资活动", "自由现金流",
                      "operating", "investing", "financing", "free cash flow"],
        "估值模型": ["收入", "增长率", "毛利率", "ebitda", "pe", "ps", "ev", "估值",
                    "wacc", "terminal", "终值"],
    }
    keywords = metric_keywords.get(sheet_type, [])
    if not keywords:
        return []
    key_rows = []
    for row in rows:
        cell0 = str(row[0]).lower() if row and row[0] else ""
        if any(k in cell0 for k in keywords):
            key_rows.append(row)
    return key_rows


def _format_row(row: list, max_cols: int) -> str:
    """格式化一行，截断过宽内容"""
    cells = []
    for i, cell in enumerate(row[:max_cols]):
        val = str(cell) if cell is not None else ""
        if len(val) > 60:
            val = val[:57] + "..."
        cells.append(val)
    return "| " + " | ".join(cells) + " |"


def load_excel_file(path: Path) -> str:
    """用 openpyxl 读取 xlsx/xls，输出结构化 markdown + 财务摘要"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        wb_data = openpyxl.load_workbook(path, data_only=True)
    except ImportError:
        return f"[缺少openpyxl: {path.name}]"
    except Exception as e:
        return f"[无法读取: {path.name}, {e}]"

    MAX_COLS = 20
    MAX_ROWS = 200
    KEY_ROW_LIMIT = 30

    parts = []
    sheet_types = {}
    all_key_metrics = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]
        if ws.max_row is None or ws.max_column is None:
            parts.append(f"## Sheet: {sheet_name}\n[空sheet]")
            continue

        total_rows = ws.max_row
        total_cols = min(ws.max_column, MAX_COLS)

        # 读取所有行数据（公式 + 缓存值）
        raw_rows = []
        for r_idx, (row, row_data) in enumerate(zip(
            ws.iter_rows(min_col=1, max_col=total_cols, values_only=False),
            ws_data.iter_rows(min_col=1, max_col=total_cols, values_only=True),
        )):
            cells = []
            for cell, data_val in zip(row, row_data):
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula = cell.value
                        if data_val is not None:
                            cells.append(f"{data_val} [公式: {formula}]")
                        else:
                            cells.append(f"[公式: {formula}]")
                    else:
                        cells.append(cell.value)
                else:
                    cells.append(None)
            raw_rows.append(cells)

        if not raw_rows:
            parts.append(f"## Sheet: {sheet_name}\n[空sheet]")
            continue

        # 分类 sheet
        header_row = raw_rows[0] if raw_rows else []
        sheet_type = _classify_sheet(sheet_name, header_row)
        sheet_types[sheet_name] = sheet_type

        # 决定显示行数
        display_rows = raw_rows[:MAX_ROWS]

        # 对大表做智能截断：保留头部、关键行、尾部
        if total_rows > MAX_ROWS:
            key_rows = _extract_key_metrics(sheet_type, raw_rows, total_cols)
            tail_rows = raw_rows[-5:] if len(raw_rows) > 5 else []
            seen_indices = set(range(min(20, len(raw_rows))))
            for kr in key_rows[:KEY_ROW_LIMIT]:
                if kr in raw_rows:
                    seen_indices.add(raw_rows.index(kr))
            for tr in tail_rows:
                if tr in raw_rows:
                    seen_indices.add(raw_rows.index(tr))
            display_rows = [raw_rows[i] for i in sorted(seen_indices) if i < len(raw_rows)]

        # 构建 markdown 表格
        col_count = max((len(r) for r in display_rows), default=0)
        header = _format_row(
            [f"Col_{i+1}" if (not header_row or i >= len(header_row) or header_row[i] is None) else header_row[i]
             for i in range(col_count)],
            col_count
        )
        separator = "|" + "|".join(["---"] * col_count) + "|"

        table_lines = [f"## Sheet: {sheet_name}  `[类型: {sheet_type}]`  `({total_rows}行 × {total_cols}列)`",
                       "", header, separator]
        for row in display_rows[1:]:
            padded = row + [None] * (col_count - len(row))
            table_lines.append(_format_row(padded, col_count))

        if total_rows > MAX_ROWS:
            table_lines.append(f"\n> ⚠️ 表格过长({total_rows}行)，已智能截断保留关键行")

        parts.append("\n".join(table_lines))

        # 提取关键指标
        if sheet_type in ("利润表", "资产负债表", "现金流量表", "估值模型"):
            key_metrics = _extract_key_metrics(sheet_type, raw_rows, total_cols)
            if key_metrics:
                all_key_metrics.append((sheet_name, sheet_type, key_metrics))

    wb.close()
    wb_data.close()

    # 生成财务摘要
    summary_parts = []
    if all_key_metrics:
        summary_parts.append("# 📊 财务数据摘要")
        summary_parts.append("")
        for sname, stype, metrics in all_key_metrics:
            summary_parts.append(f"## {sname} ({stype}) 关键指标")
            for mrow in metrics[:20]:
                label = str(mrow[0]) if mrow[0] else ""
                values = [str(v) for v in mrow[1:] if v is not None]
                if label and values:
                    summary_parts.append(f"- **{label}**: {' / '.join(values)}")
            summary_parts.append("")

    # Sheet 间关联说明
    type_names = [f"{sn}({st})" for sn, st in sheet_types.items()]
    if len(type_names) > 1:
        summary_parts.append("# 📎 Sheet 间关联")
        summary_parts.append("")
        fin_types = {sn: st for sn, st in sheet_types.items() if st != "其他"}
        if "利润表" in fin_types.values() and "现金流量表" in fin_types.values():
            pl_sheet = [sn for sn, st in fin_types.items() if st == "利润表"][0]
            cf_sheet = [sn for sn, st in fin_types.items() if st == "现金流量表"][0]
            summary_parts.append(f"- `{pl_sheet}`(利润表) 的净利润 ↔ `{cf_sheet}`(现金流量表) 的经营活动现金流应可交叉验证")
        if "利润表" in fin_types.values() and "估值模型" in fin_types.values():
            pl_sheet = [sn for sn, st in fin_types.items() if st == "利润表"][0]
            val_sheet = [sn for sn, st in fin_types.items() if st == "估值模型"][0]
            summary_parts.append(f"- `{val_sheet}`(估值模型) 的收入预测应引用 `{pl_sheet}`(利润表) 的历史营收数据")
        if "资产负债表" in fin_types.values() and "利润表" in fin_types.values():
            bs_sheet = [sn for sn, st in fin_types.items() if st == "资产负债表"][0]
            pl_sheet = [sn for sn, st in fin_types.items() if st == "利润表"][0]
            summary_parts.append(f"- `{bs_sheet}`(资产负债表) 的留存收益变动 ↔ `{pl_sheet}`(利润表) 的净利润应一致")
        summary_parts.append("")

    # 组合输出
    result = "\n\n".join(parts)
    if summary_parts:
        result = "\n".join(summary_parts) + "\n\n---\n\n" + result

    return result


def load_json_file(path: Path) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return json.dumps(data, ensure_ascii=False, indent=2)
    except Exception as e:
        return f"[无法读取: {path.name}, {e}]"


LOADERS = {
    ".txt": load_text_file, ".md": load_text_file,
    ".doc": load_docx_file, ".docx": load_docx_file,
    ".csv": load_csv_file, ".json": load_json_file,
    ".xls": load_excel_file, ".xlsx": load_excel_file,
}


# ==================== PDF处理 ====================

def detect_pdf_type(path: Path) -> str:
    """
    用 OCR 检测第2页的文字数量来判断 PDF 类型
    - 第2页 > 500 字 = Word转PDF = OCR识别
    - 第2页 <= 500 字 = PPT转PDF = 转图片
    """
    try:
        import pytesseract
        from PIL import Image
        import pymupdf

        doc = pymupdf.open(path)
        
        # 只检测第2页
        if len(doc) < 2:
            doc.close()
            return "text_based"  # 不足2页默认 OCR
        
        page = doc[1]  # 第2页（0索引）
        pix = page.get_pixmap(dpi=150)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        text = pytesseract.image_to_string(img, lang='chi_sim+eng')
        chars = len(text.strip())
        
        doc.close()

        # 判断阈值
        if chars > 500:
            return "text_based"    # Word转PDF，OCR识别
        else:
            return "image_based"   # PPT转PDF，转图片

    except Exception as e:
        print(f"  PDF类型检测失败: {e}", file=sys.stderr)
        return "text_based"  # 默认当作文字型


def pdf_to_images(path: Path, cache_dir: Path, max_pages: int = 50) -> dict:
    """PPT转PDF：直接转图片，AI直接读图"""
    try:
        import pymupdf
        print(f"  [PPT转PDF→图片] {path.name}", file=sys.stderr)

        pdf_stem = path.stem
        img_dir = cache_dir / "_images" / pdf_stem
        img_dir.mkdir(parents=True, exist_ok=True)

        doc = pymupdf.open(path)
        total_pages = min(len(doc), max_pages)

        image_paths = []
        for page_num in range(total_pages):
            page = doc[page_num]
            pix = page.get_pixmap(dpi=150)
            img_name = f"page_{page_num+1:03d}.png"
            img_path = img_dir / img_name
            pix.save(str(img_path))
            image_paths.append(str(img_path))

        doc.close()
        print(f"  已转换 {total_pages} 页 -> {img_dir}", file=sys.stderr)

        return {"images": image_paths, "total_pages": total_pages, "image_dir": str(img_dir)}
    except Exception as e:
        print(f"  转图片失败: {e}", file=sys.stderr)
        return {"images": [], "total_pages": 0, "error": str(e)}


def pdf_ocr_text(path: Path, cache_dir: Path, max_pages: int = 50) -> dict:
    """Word转PDF：用OCR识别文字"""
    try:
        import pytesseract
        from PIL import Image
        import pymupdf
        print(f"  [Word转PDF→OCR] {path.name}", file=sys.stderr)

        doc = pymupdf.open(path)
        total_pages = min(len(doc), max_pages)

        full_text = []
        for page_num in range(total_pages):
            page = doc[page_num]
            pix = page.get_pixmap(dpi=200)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            text = pytesseract.image_to_string(img, lang='chi_sim+eng')
            if text and text.strip():
                lines = [l.strip() for l in text.split('\n') if l.strip() and len(l.strip()) >= 2]
                if lines:
                    full_text.append(f"[第{page_num+1}页]\n" + '\n'.join(lines))

        doc.close()

        result = '\n\n'.join(full_text)
        
        # 保存OCR结果
        pdf_stem = path.stem
        ocr_path = cache_dir / f"{pdf_stem}_ocr.txt"
        with open(ocr_path, "w", encoding="utf-8") as f:
            f.write(result)
        
        print(f"  OCR完成 {total_pages} 页 -> {ocr_path}", file=sys.stderr)

        return {"ocr_text": result, "total_pages": total_pages, "ocr_path": str(ocr_path)}
    except Exception as e:
        print(f"  OCR失败: {e}", file=sys.stderr)
        # 尝试回退到纯图片模式
        return pdf_to_images(path, cache_dir, max_pages)


def process_pdf(path: Path, cache_dir: Path, max_pages: int = 50) -> dict:
    """自动检测PDF类型并处理"""
    pdf_type = detect_pdf_type(path)
    
    if pdf_type == "image_based":
        # PPT转PDF：直接转图片
        result = pdf_to_images(path, cache_dir, max_pages)
        return {
            "file_name": path.name,
            "file_type": ".pdf",
            "file_size": path.stat().st_size,
            "pdf_type": "image_based",
            "content": "[PPT转PDF，已转为图片，请使用image工具读取]",
            "content_length": result.get("total_pages", 0),
            "content_type": "pdf_images",
            "images": result.get("images", []),
            "image_dir": result.get("image_dir", ""),
            "file_hash": file_hash(path),
        }
    else:
        # Word转PDF：OCR识别文字
        result = pdf_ocr_text(path, cache_dir, max_pages)
        return {
            "file_name": path.name,
            "file_type": ".pdf",
            "file_size": path.stat().st_size,
            "pdf_type": "text_based",
            "content": result.get("ocr_text", ""),
            "content_length": len(result.get("ocr_text", "")),
            "content_type": "pdf_ocr",
            "ocr_path": result.get("ocr_path", ""),
            "file_hash": file_hash(path),
        }


# ==================== 缓存管理 ====================

def file_hash(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def get_cache_dir(output_dir: str, source_dir: str) -> Path:
    source_name = Path(source_dir).resolve().name
    cache_dir = Path(output_dir) / source_name
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def save_text(content: str, cache_dir: Path, file_name: str) -> str:
    stem = Path(file_name).stem
    cache_path = cache_dir / f"{stem}.txt"
    with open(cache_path, "w", encoding="utf-8") as f:
        f.write(content)
    return str(cache_path)


def load_directory(directory: str, cache_dir: Path, max_pages: int = 50) -> list:
    dir_path = Path(directory)
    if not dir_path.exists():
        print(json.dumps({"error": f"目录不存在: {directory}"}, ensure_ascii=False))
        sys.exit(1)

    documents = []
    for fp in sorted(dir_path.rglob("*")):
        if fp.is_file() and fp.suffix.lower() in SUPPORTED_EXTENSIONS:
            try:
                if fp.suffix.lower() == '.pdf':
                    doc = process_pdf(fp, cache_dir, max_pages)
                    documents.append(doc)
                else:
                    loader = LOADERS.get(fp.suffix.lower(), load_text_file)
                    content = loader(fp)
                    documents.append({
                        "file_name": fp.name,
                        "file_type": fp.suffix.lower(),
                        "file_size": fp.stat().st_size,
                        "content": content,
                        "content_length": len(content),
                        "content_type": "text",
                        "file_hash": file_hash(fp),
                    })

                print(f"  已加载: {fp.name} ({fp.stat().st_size // 1024}KB)", file=sys.stderr)
            except Exception as e:
                print(f"  加载失败: {fp.name} - {e}", file=sys.stderr)

    return documents


def main():
    parser = argparse.ArgumentParser(description="文档加载器 - PDF自动识别类型处理")
    parser.add_argument("directory", help="资料目录路径")
    parser.add_argument("--max-pages", type=int, default=50, help="PDF最大处理页数（默认50）")
    parser.add_argument("--output-dir", default="", help="缓存目录")
    args = parser.parse_args()

    if args.output_dir:
        cache_dir = get_cache_dir(args.output_dir, args.directory)
    else:
        parent = str(Path(args.directory).resolve().parent)
        cache_dir = get_cache_dir(str(Path(parent) / "_doc_cache"), args.directory)

    documents = load_directory(args.directory, cache_dir, args.max_pages)

    files = []
    for doc in documents:
        file_info = {
            "file_name": doc["file_name"],
            "file_type": doc["file_type"],
            "file_size": doc["file_size"],
            "pdf_type": doc.get("pdf_type", "N/A"),
            "content_type": doc.get("content_type", "text"),
            "file_hash": doc["file_hash"],
        }

        if doc.get("content_type") == "pdf_images":
            file_info["images"] = doc.get("images", [])
            file_info["image_dir"] = doc.get("image_dir", "")
            file_info["total_pages"] = doc.get("content_length", 0)
        elif doc.get("content_type") == "pdf_ocr":
            file_info["ocr_path"] = doc.get("ocr_path", "")
            file_info["content_length"] = doc.get("content_length", 0)
        else:
            cache_path = save_text(doc["content"], cache_dir, doc["file_name"])
            file_info["cache_path"] = cache_path
            file_info["content_length"] = doc.get("content_length", 0)

        files.append(file_info)

    manifest = {
        "source_directory": str(Path(args.directory).resolve()),
        "cache_dir": str(cache_dir),
        "extracted_at": datetime.now().isoformat(),
        "total_documents": len(files),
        "files": files,
    }
    manifest_path = cache_dir / "_manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    print(f"  缓存已保存到: {cache_dir}", file=sys.stderr)

    output = {
        "directory": args.directory,
        "cache_dir": str(cache_dir),
        "manifest_path": str(manifest_path),
        "total_documents": len(files),
        "files": files,
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()



